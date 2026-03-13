"""
Module de configuration et lecture Excel pour l'import Salesforce multi-onglets.
- Lit l'ordre des onglets depuis une feuille _Config_
- Gère les correspondances entre colonnes (source -> cible)
- Permet la réplication des IDs Salesforce entre onglets
"""

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


CONFIG_SHEET = "_Config_"
ORDER_SECTION = "OrdreOnglets"
MAPPING_SECTION = "Correspondances"
PIVOT_SECTION = "ColonnesPivot"


@dataclass
class SheetConfig:
    """Configuration d'un onglet de données."""
    name: str
    order: int
    salesforce_object: str
    id_column: Optional[str] = None  # Colonne où stocker l'ID Salesforce après import
    mode: str = "insert"  # insert | update | upsert
    upsert_external_id_field: Optional[str] = None  # Champ SF pour upsert (ex: RefExterne__c)


@dataclass
class ColumnMapping:
    """Correspondance entre une colonne source et une colonne cible."""
    source_sheet: str
    source_column: str
    target_sheet: str
    target_column: str
    target_salesforce_field: Optional[str] = None  # Nom du champ SF si différent


@dataclass
class PivotColumn:
    """Colonne pivot : utilisée pour la correspondance avec un onglet parent, mais non importée."""
    sheet: str
    column: str


@dataclass
class ImportConfig:
    """Configuration complète de l'import."""
    sheets: list[SheetConfig] = field(default_factory=list)
    mappings: list[ColumnMapping] = field(default_factory=list)
    pivot_columns: list[PivotColumn] = field(default_factory=list)
    
    def get_sheets_ordered(self) -> list[SheetConfig]:
        """Retourne les onglets triés par ordre."""
        return sorted(self.sheets, key=lambda s: s.order)
    
    def get_mappings_for_target(self, target_sheet: str) -> list[ColumnMapping]:
        """Retourne les correspondances dont la cible est l'onglet donné."""
        return [m for m in self.mappings if m.target_sheet == target_sheet]
    
    def get_pivot_columns_for_sheet(self, sheet_name: str) -> set[str]:
        """
        Colonnes à exclure de l'import pour cet onglet (pivot uniquement).
        Inclut : ColonnesPivot explicites + target_column des correspondances.
        """
        pivot = {pc.column for pc in self.pivot_columns if pc.sheet == sheet_name}
        for m in self.mappings:
            if m.target_sheet == sheet_name:
                pivot.add(m.target_column)
        return pivot


def read_excel_file(filepath: str) -> dict[str, pd.DataFrame]:
    """Lit un fichier Excel et retourne un dictionnaire {nom_onglet: DataFrame}."""
    xl = pd.ExcelFile(filepath)
    data = {}
    for sheet_name in xl.sheet_names:
        if sheet_name != CONFIG_SHEET:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
            data[sheet_name] = df
    return data


def read_config_from_excel(filepath: str) -> Optional[ImportConfig]:
    """
    Lit la configuration depuis l'onglet _Config_ du fichier Excel.
    
    Structure attendue de l'onglet _Config_ :
    
    Section OrdreOnglets (header sur une ligne) :
    | NomOnglet | Ordre | ObjetSalesforce | ColonneID |
    
    Section Correspondances (header sur une ligne) :
    | FeuilleSource | ColonneSource | FeuilleCible | ColonneCible | ChampSF |
    """
    try:
        xl = pd.ExcelFile(filepath)
        if CONFIG_SHEET not in xl.sheet_names:
            return None
        
        df_config = pd.read_excel(filepath, sheet_name=CONFIG_SHEET, header=None)
        config = ImportConfig()
        
        current_section = None
        headers = []
        
        for idx, row in df_config.iterrows():
            row_values = [str(v).strip() if pd.notna(v) else "" for v in row]
            first_cell = row_values[0] if row_values else ""
            
            if first_cell == ORDER_SECTION:
                current_section = ORDER_SECTION
                headers = []
                continue
            elif first_cell == MAPPING_SECTION:
                current_section = MAPPING_SECTION
                headers = []
                continue
            elif first_cell == PIVOT_SECTION:
                current_section = PIVOT_SECTION
                headers = []
                continue
            
            if current_section == ORDER_SECTION:
                if not headers and any("nom" in str(v).lower() or "onglet" in str(v).lower() for v in row_values):
                    headers = [h for h in row_values if h]
                # Parser les lignes de données (nom en col 1, ordre en col 2)
                if len(row_values) >= 2:
                    try:
                        name = str(row_values[1]).strip() if len(row_values) > 1 else ""
                        order_val = row_values[2] if len(row_values) > 2 else ""
                        order = int(float(order_val)) if order_val != "" and str(order_val) not in ("nan", "") else -1
                        sf_obj = str(row_values[3]).strip() if len(row_values) > 3 else ""
                        id_col = str(row_values[4]).strip() if len(row_values) > 4 and row_values[4] else None
                        mode_raw = str(row_values[5]).strip().lower() if len(row_values) > 5 and row_values[5] else "insert"
                        mode = mode_raw if mode_raw in ("insert", "update", "upsert") else "insert"
                        upsert_field = str(row_values[6]).strip() if len(row_values) > 6 and row_values[6] else None
                        if upsert_field and str(upsert_field).lower() in ("nan", ""):
                            upsert_field = None
                        # Ignorer la ligne d'en-tête (contient "onglet" ou "nom" dans la colonne nom)
                        if name and name.lower() not in ("nan", "") and "onglet" not in name.lower() and order >= 0:
                            config.sheets.append(SheetConfig(
                                name=name, order=order,
                                salesforce_object=sf_obj,
                                id_column=id_col or "Id",
                                mode=mode,
                                upsert_external_id_field=upsert_field
                            ))
                    except (ValueError, TypeError):
                        pass
            
            elif current_section == MAPPING_SECTION:
                if not headers and any("source" in str(v).lower() or "feuille" in str(v).lower() for v in row_values):
                    headers = [h for h in row_values if h]
                # Parser les lignes de correspondances (ignorer la ligne d'en-tête)
                if len(row_values) >= 4:
                    try:
                        src_sheet = str(row_values[1]).strip() if len(row_values) > 1 else ""
                        src_col = str(row_values[2]).strip() if len(row_values) > 2 else ""
                        tgt_sheet = str(row_values[3]).strip() if len(row_values) > 3 else ""
                        tgt_col = str(row_values[4]).strip() if len(row_values) > 4 else ""
                        sf_raw = row_values[5] if len(row_values) > 5 else None
                        sf_field = str(sf_raw).strip() if sf_raw and str(sf_raw).lower() != "nan" else None
                        skip_headers = ("feuille", "source", "cible", "colonne", "champ")
                        if any(h in src_sheet.lower() or h in src_col.lower() for h in skip_headers):
                            pass
                        elif all([src_sheet, src_col, tgt_sheet, tgt_col]):
                            config.mappings.append(ColumnMapping(
                                source_sheet=src_sheet,
                                source_column=src_col,
                                target_sheet=tgt_sheet,
                                target_column=tgt_col,
                                target_salesforce_field=sf_field
                            ))
                    except (ValueError, TypeError):
                        pass
            
            elif current_section == PIVOT_SECTION:
                if len(row_values) >= 2:
                    try:
                        sheet_name = str(row_values[1]).strip() if len(row_values) > 1 else ""
                        col_name = str(row_values[2]).strip() if len(row_values) > 2 else ""
                        skip = ("feuille", "colonne")
                        if sheet_name and col_name and sheet_name.lower() not in skip and col_name.lower() not in skip:
                            config.pivot_columns.append(PivotColumn(sheet=sheet_name, column=col_name))
                    except (ValueError, TypeError):
                        pass
        
        return config if config.sheets else None
    except Exception:
        return None


def add_import_columns(data: dict[str, pd.DataFrame], config: ImportConfig) -> dict[str, pd.DataFrame]:
    """
    Ajoute les colonnes nécessaires pour l'import :
    - Colonne Id : sur chaque onglet, pour recevoir l'ID Salesforce après import
    - Colonnes de référence : sur les onglets d'ordre supérieur, pour recevoir
      l'ID de l'objet des onglets d'ordre inférieur (ex: AccountId sur Contacts)
    """
    result = {}
    for name, df in data.items():
        result[name] = df.copy()

    if not config:
        return result

    # 1. Ajouter la colonne Id sur chaque onglet (sera remplie après son import)
    for sheet_config in config.sheets:
        name = sheet_config.name
        if name not in result:
            continue
        id_col = sheet_config.id_column or "Id"
        if id_col not in result[name].columns:
            result[name] = result[name].copy()
            result[name][id_col] = None

    # 2. Ajouter les colonnes de référence sur les onglets cibles (ordre supérieur)
    for mapping in config.mappings:
        target_sheet = mapping.target_sheet
        if target_sheet not in result:
            continue
        target_id_col = mapping.target_salesforce_field or f"{mapping.source_sheet}Id"
        if target_id_col not in result[target_sheet].columns:
            result[target_sheet] = result[target_sheet].copy()
            result[target_sheet][target_id_col] = None

    return result


def create_default_config(sheet_names: list[str]) -> ImportConfig:
    """Crée une configuration par défaut avec tous les onglets ordonnés 1..n."""
    config = ImportConfig()
    for i, name in enumerate(sheet_names, 1):
        config.sheets.append(SheetConfig(
            name=name, order=i,
            salesforce_object="",  # À configurer par l'utilisateur
            id_column="Id",
            mode="insert",
        ))
    return config


def apply_id_mapping(
    source_df: pd.DataFrame,
    source_key_col: str,
    source_id_col: str,
    target_df: pd.DataFrame,
    target_key_col: str,
    target_id_col: str
) -> pd.DataFrame:
    """
    Réplique les IDs de source_df vers target_df en fonction des clés correspondantes.
    source_key_col/target_key_col : colonnes de correspondance (ex: RéfExterne)
    source_id_col : colonne contenant l'ID Salesforce dans source
    target_id_col : colonne à remplir dans target
    """
    result = target_df.copy()
    if source_key_col not in source_df.columns or source_id_col not in source_df.columns:
        return result
    if target_key_col not in result.columns:
        result[target_key_col] = None
    if target_id_col not in result.columns:
        result[target_id_col] = None
    
    id_map = dict(zip(
        source_df[source_key_col].astype(str),
        source_df[source_id_col].astype(str)
    ))
    
    result[target_id_col] = result[target_key_col].astype(str).map(id_map)
    return result


def save_config_to_excel(config: ImportConfig, filepath: str) -> None:
    """
    Enregistre la configuration dans l'onglet _Config_ du fichier Excel.
    Met à jour ou crée l'onglet _Config_ selon le cas.
    """
    wb = load_workbook(filepath)
    if CONFIG_SHEET in wb.sheetnames:
        ws = wb[CONFIG_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(CONFIG_SHEET, 0)

    row_num = 1
    ws.cell(row=row_num, column=1, value=ORDER_SECTION)
    row_num += 1
    ws.cell(row=row_num, column=1, value="")
    ws.cell(row=row_num, column=2, value="NomOnglet")
    ws.cell(row=row_num, column=3, value="Ordre")
    ws.cell(row=row_num, column=4, value="ObjetSalesforce")
    ws.cell(row=row_num, column=5, value="ColonneID")
    ws.cell(row=row_num, column=6, value="Mode")
    ws.cell(row=row_num, column=7, value="ChampUpsert")
    row_num += 1
    for s in config.get_sheets_ordered():
        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value=s.name)
        ws.cell(row=row_num, column=3, value=s.order)
        ws.cell(row=row_num, column=4, value=s.salesforce_object or "")
        ws.cell(row=row_num, column=5, value=s.id_column or "Id")
        ws.cell(row=row_num, column=6, value=getattr(s, "mode", "insert") or "insert")
        ws.cell(row=row_num, column=7, value=getattr(s, "upsert_external_id_field", None) or "")
        row_num += 1
    row_num += 1

    ws.cell(row=row_num, column=1, value=MAPPING_SECTION)
    row_num += 1
    ws.cell(row=row_num, column=1, value="")
    ws.cell(row=row_num, column=2, value="FeuilleSource")
    ws.cell(row=row_num, column=3, value="ColonneSource")
    ws.cell(row=row_num, column=4, value="FeuilleCible")
    ws.cell(row=row_num, column=5, value="ColonneCible")
    ws.cell(row=row_num, column=6, value="ChampSF")
    row_num += 1
    for m in config.mappings:
        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value=m.source_sheet)
        ws.cell(row=row_num, column=3, value=m.source_column)
        ws.cell(row=row_num, column=4, value=m.target_sheet)
        ws.cell(row=row_num, column=5, value=m.target_column)
        ws.cell(row=row_num, column=6, value=m.target_salesforce_field or "")
        row_num += 1
    row_num += 1

    ws.cell(row=row_num, column=1, value=PIVOT_SECTION)
    row_num += 1
    ws.cell(row=row_num, column=1, value="")
    ws.cell(row=row_num, column=2, value="Feuille")
    ws.cell(row=row_num, column=3, value="Colonne")
    row_num += 1
    for pc in config.pivot_columns:
        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value=pc.sheet)
        ws.cell(row=row_num, column=3, value=pc.column)
        row_num += 1

    wb.save(filepath)


def save_config_to_json(config: ImportConfig, filepath: str) -> None:
    """Sauvegarde la configuration en JSON."""
    data = {
        "sheets": [
            {
                "name": s.name, "order": s.order, "salesforce_object": s.salesforce_object,
                "id_column": s.id_column, "mode": getattr(s, "mode", "insert"),
                "upsert_external_id_field": getattr(s, "upsert_external_id_field", None)
            }
            for s in config.sheets
        ],
        "mappings": [
            {"source_sheet": m.source_sheet, "source_column": m.source_column,
             "target_sheet": m.target_sheet, "target_column": m.target_column,
             "target_salesforce_field": m.target_salesforce_field}
            for m in config.mappings
        ],
        "pivot_columns": [
            {"sheet": pc.sheet, "column": pc.column}
            for pc in config.pivot_columns
        ]
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def load_config_from_json(filepath: str) -> ImportConfig:
    """Charge la configuration depuis un fichier JSON."""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    config = ImportConfig()
    _sheet_keys = {"name", "order", "salesforce_object", "id_column", "mode", "upsert_external_id_field"}
    for s in data.get("sheets", []):
        kwargs = {k: v for k, v in s.items() if k in _sheet_keys}
        config.sheets.append(SheetConfig(**kwargs))
    for m in data.get("mappings", []):
        config.mappings.append(ColumnMapping(**m))
    for pc in data.get("pivot_columns", []):
        config.pivot_columns.append(PivotColumn(**pc))
    return config
