"""
Script pour créer un fichier Excel d'exemple avec la structure attendue
pour l'import Salesforce multi-onglets.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_FILE = Path(__file__).parent / "exemple_import_salesforce.xlsx"


def create_example():
    wb = Workbook()
    ws_config = wb.active
    ws_config.title = "_Config_"

    # === SECTION 1 : Ordre des onglets (autant que nécessaire) ===
    ws_config.append(["OrdreOnglets"])
    ws_config.append(["", "NomOnglet", "Ordre", "ObjetSalesforce", "ColonneID"])
    ws_config.append(["", "Comptes", 1, "Account", "Id"])
    ws_config.append(["", "Contacts", 2, "Contact", "Id"])
    ws_config.append(["", "Opportunites", 3, "Opportunity", "Id"])
    ws_config.append([])

    # === SECTION 2 : Correspondances (plusieurs liens possibles) ===
    ws_config.append(["Correspondances"])
    ws_config.append(["", "FeuilleSource", "ColonneSource", "FeuilleCible", "ColonneCible", "ChampSF"])
    ws_config.append(["", "Comptes", "RefExterne", "Contacts", "RefCompte", "AccountId"])
    ws_config.append(["", "Comptes", "RefExterne", "Opportunites", "RefCompte", "AccountId"])
    ws_config.append([])

    # === SECTION 3 : Colonnes pivot ===
    ws_config.append(["ColonnesPivot"])
    ws_config.append(["", "Feuille", "Colonne"])
    ws_config.append(["", "Contacts", "RefCompte"])
    ws_config.append(["", "Contacts", "NotesImport"])
    ws_config.append(["", "Opportunites", "RefCompte"])

    # === Feuille Comptes (ordre 1) ===
    ws_comptes = wb.create_sheet("Comptes")
    comptes = pd.DataFrame({
        "RefExterne": ["C001", "C002", "C003"],
        "Name": ["Entreprise Alpha", "Entreprise Beta", "Entreprise Gamma"],
        "BillingCity": ["Paris", "Lyon", "Marseille"],
    })
    for r in dataframe_to_rows(comptes, index=False, header=True):
        ws_comptes.append(r)

    # === Feuille Contacts (ordre 2) ===
    ws_contacts = wb.create_sheet("Contacts")
    contacts = pd.DataFrame({
        "RefCompte": ["C001", "C001", "C002"],
        "NotesImport": ["Note 1", "Note 2", "Note 3"],
        "LastName": ["Dupont", "Martin", "Bernard"],
        "FirstName": ["Jean", "Marie", "Pierre"],
        "Email": ["jean.dupont@alpha.com", "marie.martin@alpha.com", "pierre.bernard@beta.com"],
    })
    for r in dataframe_to_rows(contacts, index=False, header=True):
        ws_contacts.append(r)

    # === Feuille Opportunites (ordre 3) ===
    ws_opp = wb.create_sheet("Opportunites")
    opportunites = pd.DataFrame({
        "RefCompte": ["C001", "C002", "C001"],
        "Name": ["Deal Alpha", "Deal Beta", "Deal Gamma"],
        "StageName": ["Prospecting", "Qualification", "Prospecting"],
        "CloseDate": ["2025-06-30", "2025-07-15", "2025-08-01"],
    })
    for r in dataframe_to_rows(opportunites, index=False, header=True):
        ws_opp.append(r)

    wb.save(OUTPUT_FILE)
    print(f"Fichier créé : {OUTPUT_FILE}")
    print("\nStructure : 3 onglets (Comptes → Contacts → Opportunites)")
    print("  - OrdreOnglets : définir autant d'onglets que nécessaire (ordre 1, 2, 3, ...)")
    print("  - Correspondances : plusieurs liens possibles (ex: Comptes→Contacts, Comptes→Opportunites)")
    print("  - ColonnesPivot : RefCompte, NotesImport (non importées)")


if __name__ == "__main__":
    create_example()
