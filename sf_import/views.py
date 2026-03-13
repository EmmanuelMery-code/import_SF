import io
import json
import uuid
from urllib.parse import urlencode

import pandas as pd
from django.conf import settings
from django.contrib import messages
from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from excel_import_config import (
    read_excel_file,
    read_config_from_excel,
    create_default_config,
    add_import_columns,
    save_config_to_excel,
    ImportConfig,
    SheetConfig,
    ColumnMapping,
    PivotColumn,
    CONFIG_SHEET,
)
from salesforce_importer import (
    prepare_records_for_salesforce,
    import_to_salesforce,
    replicate_ids_to_next_sheets,
)
from .salesforce_oauth import create_salesforce_client, get_org_info


def _get_sf_client(request):
    """Récupère le client Salesforce depuis la session."""
    creds = request.session.get("salesforce_credentials")
    if not creds:
        return None
    try:
        return create_salesforce_client(
            instance_url=creds["instance_url"],
            session_id=creds["session_id"],
        )
    except Exception:
        return None


@require_GET
def create_import_file(request):
    """Page pour créer un fichier Excel d'import à partir des objets Salesforce."""
    sf_client = _get_sf_client(request)
    if not sf_client:
        messages.warning(request, "Connectez-vous à Salesforce pour accéder à cette page.")
        return redirect("sf_login")

    sobjects = []
    try:
        desc = sf_client.describe()
        for obj in desc.get("sobjects", []):
            if obj.get("createable") and obj.get("name") and not obj.get("deprecatedAndHidden"):
                sobjects.append({
                    "name": obj["name"],
                    "label": obj.get("label", obj["name"]),
                    "custom": obj.get("custom", False),
                })
        sobjects.sort(key=lambda x: (not x["custom"], x["label"].lower()))
    except Exception as e:
        sobjects = []
        messages.error(request, f"Erreur lors de la récupération des objets : {e}")

    context = {
        "sobjects": sobjects,
        "connected": True,
    }
    return render(request, "sf_import/create_import_file.html", context)


@require_POST
def create_import_file_generate(request):
    """Génère et retourne le fichier Excel pour import."""
    sf_client = _get_sf_client(request)
    if not sf_client:
        return redirect("sf_login")

    objects = request.POST.getlist("objects")
    if not objects:
        messages.error(request, "Sélectionnez au moins un objet.")
        return redirect("create_import_file")

    try:
        wb = Workbook()
        wb.remove(wb.active)

        config = ImportConfig()
        for idx, obj_name in enumerate(objects, 1):
            try:
                sobj = getattr(sf_client, obj_name)
                desc = sobj.describe()
            except Exception:
                continue

            fields = []
            for f in desc.get("fields", []):
                if not f.get("createable") or f.get("name") in ("Id",):
                    continue
                if f.get("type") == "address" or f.get("compoundFieldName"):
                    continue
                fields.append(f)

            sheet_name = obj_name[:31]
            ws = wb.create_sheet(sheet_name, idx)
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col_idx, value=field["name"])
                parts = [
                    f"Label: {field.get('label', '')}",
                    f"Type: {field.get('type', '')}",
                ]
                if field.get("length"):
                    parts.append(f"Longueur max: {field['length']}")
                if field.get("picklistValues"):
                    opts = [v["value"] for v in field["picklistValues"] if v.get("active")]
                    if opts:
                        parts.append(f"Valeurs: {', '.join(opts[:10])}{'...' if len(opts) > 10 else ''}")
                if field.get("referenceTo"):
                    parts.append(f"Référence: {', '.join(field['referenceTo'])}")
                if not field.get("nillable"):
                    parts.append("Requis")
                cell.comment = Comment("\n".join(parts), "Import SF")

            config.sheets.append(SheetConfig(
                name=sheet_name,
                order=idx,
                salesforce_object=obj_name,
                id_column="Id",
                mode="insert",
            ))

        ws_config = wb.create_sheet(CONFIG_SHEET, 0)
        _write_config_sheet(ws_config, config)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        resp = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="import_salesforce.xlsx"'
        return resp
    except Exception as e:
        messages.error(request, f"Erreur : {e}")
        return redirect("create_import_file")


def _write_config_sheet(ws, config):
    """Écrit l'onglet _Config_ selon la structure attendue."""
    row = 1
    ws.cell(row=row, column=1, value="OrdreOnglets")
    row += 1
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="NomOnglet")
    ws.cell(row=row, column=3, value="Ordre")
    ws.cell(row=row, column=4, value="ObjetSalesforce")
    ws.cell(row=row, column=5, value="ColonneID")
    ws.cell(row=row, column=6, value="Mode")
    ws.cell(row=row, column=7, value="ChampUpsert")
    row += 1
    for s in config.get_sheets_ordered():
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value=s.name)
        ws.cell(row=row, column=3, value=s.order)
        ws.cell(row=row, column=4, value=s.salesforce_object or "")
        ws.cell(row=row, column=5, value=s.id_column or "Id")
        ws.cell(row=row, column=6, value=getattr(s, "mode", "insert") or "insert")
        ws.cell(row=row, column=7, value=getattr(s, "upsert_external_id_field", None) or "")
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Correspondances")
    row += 1
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="FeuilleSource")
    ws.cell(row=row, column=3, value="ColonneSource")
    ws.cell(row=row, column=4, value="FeuilleCible")
    ws.cell(row=row, column=5, value="ColonneCible")
    ws.cell(row=row, column=6, value="ChampSF")
    row += 1
    ws.cell(row=row, column=1, value="ColonnesPivot")
    row += 1
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="Feuille")
    ws.cell(row=row, column=3, value="Colonne")


def index(request):
    """Page d'accueil avec upload de fichier."""
    if request.session.get("preferred_view") == "explorer" and request.GET.get("view") != "classic":
        return redirect("explorer")
    request.session["preferred_view"] = "index"
    context = {
        "connected": bool(request.session.get("salesforce_credentials")),
        "current_file_name": request.session.get("current_file_name"),
    }
    return render(request, "sf_import/index.html", context)


def _build_sheet_tree(config):
    """
    Construit l'arborescence des onglets selon les dépendances (correspondances).
    target_sheet dépend de source_sheet → target est enfant de source.
    """
    if not config or not config.sheets:
        return []
    sheets_ordered = config.get_sheets_ordered()
    sheet_by_name = {s.name: s for s in sheets_ordered}
    # Parents : pour chaque target, son parent = source du mapping (le premier par ordre si plusieurs)
    targets_from = {}  # target -> list of (source_order, source_name)
    for m in config.mappings:
        if m.target_sheet in sheet_by_name and m.source_sheet in sheet_by_name:
            if m.target_sheet not in targets_from:
                targets_from[m.target_sheet] = []
            src_order = next((s.order for s in sheets_ordered if s.name == m.source_sheet), 999)
            targets_from[m.target_sheet].append((src_order, m.source_sheet))
    # Pour chaque target, parent = source avec le plus petit ordre
    parent_of = {}
    for target, sources in targets_from.items():
        sources.sort(key=lambda x: x[0])
        parent_of[target] = sources[0][1]
    # Enfants directs
    children_of = {s.name: [] for s in sheets_ordered}
    for target, parent in parent_of.items():
        children_of[parent].append(target)
    # Tri des enfants par ordre d'import
    order_idx = {s.name: i for i, s in enumerate(sheets_ordered)}
    for name in children_of:
        children_of[name].sort(key=lambda x: order_idx.get(x, 999))
    # Racines = onglets qui ne sont jamais target
    roots = [s.name for s in sheets_ordered if s.name not in parent_of]
    roots.sort(key=lambda x: order_idx.get(x, 999))

    def build_node(name):
        s = sheet_by_name.get(name)
        if not s:
            return None
        children = []
        for c in children_of.get(name, []):
            n = build_node(c)
            if n:
                children.append(n)
        return {"name": name, "order": s.order, "object": s.salesforce_object, "children": children}

    return [build_node(r) for r in roots]


@require_GET
def explorer(request):
    """Mode explorateur : arborescence à gauche, contenu à droite."""
    request.session["preferred_view"] = "explorer"
    creds = request.session.get("salesforce_credentials")
    connected = bool(creds)
    file_id = request.session.get("current_file_id")
    file_name = request.session.get("current_file_name", "")

    tree_nodes = []
    has_file = False
    config = None
    sheet_names = []
    columns_by_sheet = {}
    config_sheet_rows = []
    config_summary = []
    sheets_data = []
    tree_json = "[]"
    common_sf_objects = ["Account", "Contact", "Lead", "Opportunity", "CustomObject__c"]
    sheet_names_json = "[]"
    columns_by_sheet_json = "{}"
    sobjects = []

    if connected:
        sf_client = _get_sf_client(request)
        if sf_client:
            try:
                desc = sf_client.describe()
                for obj in desc.get("sobjects", []):
                    if obj.get("createable") and obj.get("name") and not obj.get("deprecatedAndHidden"):
                        sobjects.append({
                            "name": obj["name"],
                            "label": obj.get("label", obj["name"]),
                            "custom": obj.get("custom", False),
                        })
                sobjects.sort(key=lambda x: (not x["custom"], x["label"].lower()))
            except Exception:
                pass

    if file_id:
        filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
        if filepath.exists():
            has_file = True
            try:
                data = read_excel_file(str(filepath))
                config = read_config_from_excel(str(filepath))
                if not config:
                    config = create_default_config(list(data.keys()))
                data = add_import_columns(data, config)
                sheets_ordered = config.get_sheets_ordered()
                sheet_names = [n for n in data.keys() if n != "_Config_"]
                columns_by_sheet = {n: list(df.columns) for n, df in data.items()}
                config_summary = _config_summary(config)
                tree_nodes = _build_sheet_tree(config)
                tree_json = json.dumps(tree_nodes, ensure_ascii=False)

                for sheet_config in sheets_ordered:
                    name = sheet_config.name
                    if name not in data:
                        continue
                    df = data[name]
                    cols = list(df.columns)
                    rows = []
                    for _, row in df.head(200).iterrows():
                        rows.append([str(v)[:80] if pd.notna(v) else "" for v in row])
                    sheets_data.append({
                        "name": name,
                        "order": sheet_config.order,
                        "object": sheet_config.salesforce_object,
                        "mode": getattr(sheet_config, "mode", "insert") or "insert",
                        "upsert_external_id_field": getattr(sheet_config, "upsert_external_id_field", None) or "",
                        "columns": cols,
                        "rows": rows,
                        "total_rows": len(df),
                    })

                for s in config.get_sheets_ordered():
                    config_sheet_rows.append({
                        "name": s.name,
                        "order": s.order,
                        "salesforce_object": s.salesforce_object or "",
                        "id_column": s.id_column or "Id",
                        "mode": getattr(s, "mode", "insert") or "insert",
                        "upsert_field": getattr(s, "upsert_external_id_field", None) or "",
                        "columns": columns_by_sheet.get(s.name, []),
                    })
                sheet_names_json = json.dumps(sheet_names, ensure_ascii=False)
                columns_by_sheet_json = json.dumps(columns_by_sheet, ensure_ascii=False)
            except Exception:
                has_file = False

    sobjects_json = json.dumps([{"name": o["name"], "label": o["label"]} for o in sobjects], ensure_ascii=False)

    config_graph_data = {
        "sheets": [
            {
                "name": r["name"],
                "object": r["salesforce_object"],
                "order": r["order"],
                "columns": r.get("columns", []),
                "id_column": r.get("id_column", "Id"),
                "mode": r.get("mode", "insert"),
                "upsert_field": r.get("upsert_field") or "",
            }
            for r in config_sheet_rows
        ],
        "mappings": [
            {
                "source_sheet": m.source_sheet,
                "source_column": m.source_column,
                "target_sheet": m.target_sheet,
                "target_column": m.target_column,
                "target_salesforce_field": m.target_salesforce_field or "",
            }
            for m in (config.mappings if config else [])
        ],
        "pivot_columns": [
            {"sheet": p.sheet, "column": p.column}
            for p in (config.pivot_columns if config else [])
        ],
    }
    config_graph_json = json.dumps(config_graph_data, ensure_ascii=False)

    context = {
        "connected": connected,
        "file_name": file_name,
        "has_file": has_file,
        "sobjects": sobjects,
        "sobjects_json": sobjects_json,
        "config": config,
        "tree_nodes": tree_nodes,
        "tree_json": tree_json,
        "config_summary": config_summary,
        "sheets_data": sheets_data,
        "sheet_names": sheet_names,
        "columns_by_sheet": columns_by_sheet,
        "config_sheet_rows": config_sheet_rows,
        "common_sf_objects": common_sf_objects,
        "sheet_names_json": sheet_names_json,
        "columns_by_sheet_json": columns_by_sheet_json,
        "config_graph_json": config_graph_json,
    }
    return render(request, "sf_import/explorer.html", context)


@require_http_methods(["GET", "POST"])
def upload(request):
    """Upload du fichier Excel et redirection vers la prévisualisation."""
    if request.method == "GET":
        return redirect("index")

    file = request.FILES.get("excel_file")
    if not file or not file.name.lower().endswith((".xlsx", ".xls")):
        messages.error(request, "Veuillez sélectionner un fichier Excel (.xlsx ou .xls)")
        return redirect("index")

    settings.MEDIA_ROOT.mkdir(parents=True, exist_ok=True)
    file_id = str(uuid.uuid4())
    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    with open(filepath, "wb") as f:
        for chunk in file.chunks():
            f.write(chunk)

    request.session["current_file_id"] = file_id
    request.session["current_file_name"] = file.name
    return redirect("preview")


@require_http_methods(["GET", "POST"])
def explorer_upload(request):
    """Upload du fichier Excel (mode explorateur) → redirection vers /explorer/."""
    if request.method == "GET":
        return redirect("explorer")
    file = request.FILES.get("excel_file")
    if not file or not file.name.lower().endswith((".xlsx", ".xls")):
        messages.error(request, "Veuillez sélectionner un fichier Excel (.xlsx ou .xls)")
        return redirect("explorer")
    settings.MEDIA_ROOT.mkdir(parents=True, exist_ok=True)
    file_id = str(uuid.uuid4())
    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    with open(filepath, "wb") as f:
        for chunk in file.chunks():
            f.write(chunk)
    request.session["current_file_id"] = file_id
    request.session["current_file_name"] = file.name
    return redirect("explorer")


@xframe_options_sameorigin
def preview(request):
    """Prévisualisation des onglets Excel."""
    file_id = request.session.get("current_file_id")
    if not file_id:
        messages.warning(request, "Aucun fichier chargé. Veuillez en uploader un.")
        return redirect("index")

    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    if not filepath.exists():
        messages.error(request, "Fichier introuvable. Veuillez le recharger.")
        del request.session["current_file_id"]
        return redirect("index")

    try:
        data = read_excel_file(str(filepath))
        config = read_config_from_excel(str(filepath))
        if not config:
            config = create_default_config(list(data.keys()))
        data = add_import_columns(data, config)
        sheets_ordered = config.get_sheets_ordered()
    except Exception as e:
        messages.error(request, f"Erreur de lecture : {e}")
        return redirect("index")

    sheets_data = []
    for sheet_config in sheets_ordered:
        name = sheet_config.name
        if name not in data:
            continue
        df = data[name]
        cols = list(df.columns)
        rows = []
        for _, row in df.head(100).iterrows():
            rows.append([str(v)[:50] if pd.notna(v) else "" for v in row])
        sheets_data.append({
            "name": name,
            "order": sheet_config.order,
            "object": sheet_config.salesforce_object,
            "mode": getattr(sheet_config, "mode", "insert") or "insert",
            "upsert_external_id_field": getattr(sheet_config, "upsert_external_id_field", None) or "",
            "columns": cols,
            "rows": rows,
            "total_rows": len(df),
        })

    # Données pour l'éditeur de configuration
    sheet_names = [n for n in data.keys() if n != "_Config_"]
    columns_by_sheet = {n: list(df.columns) for n, df in data.items()}
    sheet_names_json = json.dumps(sheet_names, ensure_ascii=False)
    columns_by_sheet_json = json.dumps(columns_by_sheet, ensure_ascii=False)
    common_sf_objects = ["Account", "Contact", "Lead", "Opportunity", "CustomObject__c"]
    sobjects = []
    if request.session.get("salesforce_credentials"):
        sf_client = _get_sf_client(request)
        if sf_client:
            try:
                desc = sf_client.describe()
                for obj in desc.get("sobjects", []):
                    if obj.get("createable") and obj.get("name") and not obj.get("deprecatedAndHidden"):
                        sobjects.append({"name": obj["name"], "label": obj.get("label", obj["name"]), "custom": obj.get("custom", False)})
                sobjects.sort(key=lambda x: (not x.get("custom", False), x["label"].lower()))
            except Exception:
                pass
    # Lignes de config par onglet pour le formulaire
    config_sheet_rows = []
    for s in config.get_sheets_ordered():
        config_sheet_rows.append({
            "name": s.name,
            "order": s.order,
            "salesforce_object": s.salesforce_object or "",
            "id_column": s.id_column or "Id",
            "mode": getattr(s, "mode", "insert") or "insert",
            "upsert_field": getattr(s, "upsert_external_id_field", None) or "",
            "columns": columns_by_sheet.get(s.name, []),
        })

    context = {
        "file_name": request.session.get("current_file_name", "Fichier"),
        "sheets": sheets_data,
        "config_summary": _config_summary(config),
        "connected": bool(request.session.get("salesforce_credentials")),
        "config": config,
        "sheet_names": sheet_names,
        "columns_by_sheet": columns_by_sheet,
        "sheet_names_json": sheet_names_json,
        "columns_by_sheet_json": columns_by_sheet_json,
        "config_sheet_rows": config_sheet_rows,
        "common_sf_objects": common_sf_objects,
        "sobjects": sobjects,
    }
    return render(request, "sf_import/preview.html", context)


def _config_summary(config):
    """Résumé texte de la configuration."""
    if not config:
        return []
    lines = ["=== ORDRE ==="]
    for s in config.get_sheets_ordered():
        mode = getattr(s, "mode", "insert") or "insert"
        upsert = getattr(s, "upsert_external_id_field", None) or ""
        extra = f" [mode: {mode}" + (f", champ upsert: {upsert}" if upsert else "") + "]"
        lines.append(f"  {s.order}. {s.name} → {s.salesforce_object or '(non défini)'}{extra}")
    lines.append("")
    lines.append("=== CORRESPONDANCES ===")
    for m in config.mappings:
        lines.append(f"  {m.source_sheet}.{m.source_column} → {m.target_sheet}.{m.target_column} ({m.target_salesforce_field})")
    if config.pivot_columns:
        lines.append("")
        lines.append("=== COLONNES PIVOT (non importées) ===")
        for pc in config.pivot_columns:
            lines.append(f"  {pc.sheet}.{pc.column}")
    return lines


def sf_login(request):
    """Page de connexion Salesforce : web login + récupération du session ID."""
    context = {"connected": bool(request.session.get("salesforce_credentials"))}
    return render(request, "sf_import/sf_login.html", context)


def sf_login_submit(request):
    """Enregistre le session ID et l'instance URL (formulaire manuel)."""
    if request.method != "POST":
        return redirect("sf_login")
    instance_url = request.POST.get("instance_url", "").strip()
    session_id = request.POST.get("session_id", "").replace("\r", "").replace("\n", " ").strip()
    if not instance_url or not session_id:
        messages.error(request, "Veuillez renseigner l'URL d'instance et le Session ID.")
        next_url = request.POST.get("next", "sf_login")
        return redirect(next_url)
    if not instance_url.startswith("http"):
        instance_url = "https://" + instance_url
    try:
        sf = create_salesforce_client(instance_url, session_id)
        org_info = get_org_info(sf, instance_url, session_id=session_id)
        request.session["salesforce_credentials"] = {
            "instance_url": instance_url,
            "session_id": session_id,
            "org_name": org_info.get("name", ""),
            "org_id": org_info.get("id", ""),
        }
    except Exception as e:
        messages.error(request, f"Session invalide ou expirée : {e}")
        next_url = request.POST.get("next", "sf_login")
        return redirect(next_url)
    messages.success(request, "Connexion Salesforce réussie.")
    next_url = request.POST.get("next", "index")
    return redirect(next_url)


def sf_callback(request):
    """
    Callback appelé par le bookmarklet après connexion web à Salesforce.
    Reçoit sid et instance en paramètres GET.
    Note : Le cookie sid en Lightning peut ne pas fonctionner pour l'API.
    """
    session_id = request.GET.get("sid", "").replace("\r", "").replace("\n", " ").strip()
    instance = request.GET.get("instance", "").strip()
    if not session_id or not instance:
        messages.error(request, "Session ID ou instance manquant.")
        return redirect("sf_login")
    if not instance.startswith("http"):
        instance = "https://" + instance
    try:
        sf = create_salesforce_client(instance, session_id)
        org_info = get_org_info(sf, instance, session_id=session_id)
        request.session["salesforce_credentials"] = {
            "instance_url": instance,
            "session_id": session_id,
            "org_name": org_info.get("name", ""),
            "org_id": org_info.get("id", ""),
        }
    except Exception as e:
        messages.error(request, f"Session invalide : {e}")
        return redirect("sf_login")
    messages.success(request, "Connexion Salesforce réussie.")
    next_url = request.GET.get("next", "index")
    return redirect(next_url)


def sf_logout(request):
    """Déconnexion Salesforce."""
    if "salesforce_credentials" in request.session:
        del request.session["salesforce_credentials"]
    messages.info(request, "Déconnecté de Salesforce.")
    view_name = request.session.get("preferred_view", "index")
    return redirect(view_name if view_name == "explorer" else "index")


@require_POST
def save_config(request):
    """
    Enregistre la configuration dans le fichier Excel.
    Accepte POST JSON : { sheets: [...], mappings: [...], pivot_columns: [...] }
    """
    file_id = request.session.get("current_file_id")
    if not file_id:
        return JsonResponse({"success": False, "error": "Aucun fichier chargé"})

    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    if not filepath.exists():
        return JsonResponse({"success": False, "error": "Fichier introuvable"})

    try:
        if request.content_type and "application/json" in request.content_type:
            data = json.loads(request.body)
        else:
            data = json.loads(request.POST.get("config_json", "{}"))
    except json.JSONDecodeError as e:
        return JsonResponse({"success": False, "error": f"Configuration invalide : {e}"})

    sheets_data = data.get("sheets", [])
    mappings_data = data.get("mappings", [])
    pivot_data = data.get("pivot_columns", [])

    # Charger les données Excel pour validation
    try:
        excel_data = read_excel_file(str(filepath))
        sheet_names = [n for n in excel_data.keys() if n != "_Config_"]
        columns_by_sheet = {
            n: [str(c).strip() for c in df.columns if c is not None]
            for n, df in excel_data.items()
        }
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erreur lecture Excel : {e}"})

    config = ImportConfig()

    # Parser les sheets
    seen_orders = set()
    for s in sheets_data:
        name = str(s.get("name", "")).strip()
        if not name:
            continue
        if name not in sheet_names:
            return JsonResponse({
                "success": False,
                "error": f"Onglet « {name} » inexistant dans le fichier Excel.",
            })
        try:
            order = int(s.get("order", 0))
        except (ValueError, TypeError):
            order = 0
        if order in seen_orders:
            return JsonResponse({
                "success": False,
                "error": f"Ordre {order} en double pour l'onglet « {name} ».",
            })
        seen_orders.add(order)
        sf_obj = str(s.get("salesforce_object", "")).strip()
        id_col = str(s.get("id_column", "Id")).strip() or "Id"
        mode = str(s.get("mode", "insert")).strip().lower()
        if mode not in ("insert", "update", "upsert"):
            mode = "insert"
        upsert_f = s.get("upsert_external_id_field")
        upsert_f = str(upsert_f).strip() if upsert_f else None
        config.sheets.append(SheetConfig(
            name=name, order=order, salesforce_object=sf_obj,
            id_column=id_col, mode=mode, upsert_external_id_field=upsert_f,
        ))

    if not config.sheets:
        return JsonResponse({"success": False, "error": "Aucun onglet configuré."})

    # Parser les mappings
    sheet_names_set = {s.name for s in config.sheets}
    for m in mappings_data:
        src = str(m.get("source_sheet", "")).strip()
        src_col = str(m.get("source_column", "")).strip()
        tgt = str(m.get("target_sheet", "")).strip()
        tgt_col = str(m.get("target_column", "")).strip()
        if not all([src, src_col, tgt, tgt_col]):
            continue
        if src not in sheet_names_set or tgt not in sheet_names_set:
            return JsonResponse({
                "success": False,
                "error": f"Correspondance invalide : onglet source ou cible inconnu.",
            })
        src_cols = columns_by_sheet.get(src, [])
        tgt_cols = columns_by_sheet.get(tgt, [])
        if src_col not in src_cols:
            src_matched = next((c for c in src_cols if str(c).strip() == src_col), None)
            if src_matched is not None:
                src_col = src_matched
            else:
                return JsonResponse({
                    "success": False,
                    "error": f"Colonne « {src_col} » absente dans l'onglet « {src} ».",
                })
        if tgt_col not in tgt_cols:
            tgt_matched = next((c for c in tgt_cols if str(c).strip() == tgt_col), None)
            if tgt_matched is not None:
                tgt_col = tgt_matched
            else:
                return JsonResponse({
                    "success": False,
                    "error": f"Colonne « {tgt_col} » absente dans l'onglet « {tgt} ».",
                })
        sf_field = m.get("target_salesforce_field")
        sf_field = str(sf_field).strip() if sf_field else None
        config.mappings.append(ColumnMapping(
            source_sheet=src, source_column=src_col,
            target_sheet=tgt, target_column=tgt_col,
            target_salesforce_field=sf_field,
        ))

    # Parser les pivot columns
    for p in pivot_data:
        sh = str(p.get("sheet", "")).strip()
        col = str(p.get("column", "")).strip()
        if not sh or not col:
            continue
        if sh not in sheet_names_set:
            return JsonResponse({
                "success": False,
                "error": f"Onglet « {sh} » inconnu pour la colonne pivot.",
            })
        cols = columns_by_sheet.get(sh, [])
        if col not in cols:
            matched = next((c for c in cols if str(c).strip() == col), None)
            if matched is not None:
                col = matched
            else:
                return JsonResponse({
                    "success": False,
                    "error": f"Colonne « {col} » absente dans l'onglet « {sh} ».",
                })
        config.pivot_columns.append(PivotColumn(sheet=sh, column=col))

    try:
        save_config_to_excel(config, str(filepath))
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": True, "message": "Configuration enregistrée dans l'Excel."})


@require_GET
def api_sobjects(request):
    """API : liste des objets Salesforce créables (pour sélection dans config)."""
    sf_client = _get_sf_client(request)
    if not sf_client:
        return JsonResponse({"sobjects": []})
    sobjects = []
    try:
        desc = sf_client.describe()
        for obj in desc.get("sobjects", []):
            if obj.get("createable") and obj.get("name") and not obj.get("deprecatedAndHidden"):
                sobjects.append({
                    "name": obj["name"],
                    "label": obj.get("label", obj["name"]),
                    "custom": obj.get("custom", False),
                })
        sobjects.sort(key=lambda x: (not x.get("custom", False), x["label"].lower()))
    except Exception:
        pass
    return JsonResponse({"sobjects": sobjects})


@require_POST
def add_sheet(request):
    """
    Ajoute un nouvel onglet au fichier Excel.
    POST JSON: { sheet_name, salesforce_object, parent_sheet?, source_column?, target_column?, target_sf_field? }
    Si parent_sheet est fourni, ajoute aussi la correspondance parent->enfant.
    """
    file_id = request.session.get("current_file_id")
    if not file_id:
        return JsonResponse({"success": False, "error": "Aucun fichier chargé"})
    sf_client = _get_sf_client(request)
    if not sf_client:
        return JsonResponse({"success": False, "error": "Connectez-vous à Salesforce pour ajouter un onglet"})
    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    if not filepath.exists():
        return JsonResponse({"success": False, "error": "Fichier introuvable"})
    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON invalide"})
    sheet_name = str(data.get("sheet_name", "")).strip() or str(data.get("salesforce_object", "")).strip()
    sf_object = str(data.get("salesforce_object", "")).strip()
    parent_sheet = str(data.get("parent_sheet", "")).strip() or None
    if not sheet_name or not sf_object:
        return JsonResponse({"success": False, "error": "Nom d'onglet et objet Salesforce requis"})
    sheet_name = sheet_name[:31]
    try:
        config = read_config_from_excel(str(filepath))
        if not config:
            excel_data = read_excel_file(str(filepath))
            config = create_default_config(list(excel_data.keys()))
        existing_names = {s.name for s in config.sheets}
        if sheet_name in existing_names:
            return JsonResponse({"success": False, "error": f"L'onglet « {sheet_name} » existe déjà"})
        max_order = max((s.order for s in config.sheets), default=0)
        new_order = max_order + 1
        try:
            sobj = getattr(sf_client, sf_object)
            desc = sobj.describe()
        except Exception as e:
            return JsonResponse({"success": False, "error": f"Objet Salesforce inconnu : {e}"})
        headers = []
        lookup_field_for_parent = None
        parent_obj = None
        if parent_sheet:
            parent_cfg = next((s for s in config.sheets if s.name == parent_sheet), None)
            if parent_cfg:
                parent_obj = parent_cfg.salesforce_object
        fields_meta = []
        for f in desc.get("fields", []):
            if not f.get("createable") or f.get("name") in ("Id",) or f.get("compoundFieldName") or f.get("type") == "address":
                continue
            headers.append(f["name"])
            fields_meta.append(f)
            if parent_obj and f.get("referenceTo") and parent_obj in f.get("referenceTo", []):
                lookup_field_for_parent = f["name"]
        wb = load_workbook(str(filepath))
        if sheet_name in wb.sheetnames:
            return JsonResponse({"success": False, "error": f"L'onglet « {sheet_name} » existe déjà"})
        ws = wb.create_sheet(sheet_name[:31], len(wb.sheetnames))
        for col_idx, field in enumerate(fields_meta, 1):
            cell = ws.cell(row=1, column=col_idx, value=field["name"])
            parts = [
                f"Label: {field.get('label', '')}",
                f"Type: {field.get('type', '')}",
            ]
            if field.get("length"):
                parts.append(f"Longueur max: {field['length']}")
            if field.get("picklistValues"):
                opts = [v["value"] for v in field["picklistValues"] if v.get("active")]
                if opts:
                    parts.append(f"Valeurs: {', '.join(opts[:10])}{'...' if len(opts) > 10 else ''}")
            if field.get("referenceTo"):
                parts.append(f"Référence: {', '.join(field['referenceTo'])}")
            if not field.get("nillable"):
                parts.append("Requis")
            cell.comment = Comment("\n".join(parts), "Import SF")
        config.sheets.append(SheetConfig(
            name=sheet_name, order=new_order, salesforce_object=sf_object,
            id_column="Id", mode="insert",
        ))
        if parent_sheet and parent_sheet in existing_names:
            parent_cfg = next((s for s in config.sheets if s.name == parent_sheet), None)
            src_col = str(data.get("source_column") or (parent_cfg.id_column if parent_cfg else "Id")).strip()
            tgt_col = str(data.get("target_column") or lookup_field_for_parent or "AccountId").strip()
            if tgt_col not in headers:
                headers.append(tgt_col)
                cell = ws.cell(row=1, column=len(headers), value=tgt_col)
                f_tgt = next((x for x in desc.get("fields", []) if x.get("name") == tgt_col), None)
                if f_tgt:
                    parts = [f"Label: {f_tgt.get('label', '')}", f"Type: {f_tgt.get('type', '')}"]
                    if f_tgt.get("length"):
                        parts.append(f"Longueur max: {f_tgt['length']}")
                    if f_tgt.get("referenceTo"):
                        parts.append(f"Référence: {', '.join(f_tgt['referenceTo'])}")
                    if not f_tgt.get("nillable"):
                        parts.append("Requis")
                    cell.comment = Comment("\n".join(parts), "Import SF")
            sf_field = str(data.get("target_sf_field") or tgt_col).strip()
            config.mappings.append(ColumnMapping(
                source_sheet=parent_sheet, source_column=src_col,
                target_sheet=sheet_name, target_column=tgt_col,
                target_salesforce_field=sf_field,
            ))
        wb.save(str(filepath))
        save_config_to_excel(config, str(filepath))
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": True, "sheet_name": sheet_name, "order": new_order})


@require_POST
def rename_sheet(request):
    """
    Renomme un onglet dans le fichier Excel.
    POST JSON: { old_name, new_name }
    Met à jour l'onglet, la config (OrdreOnglets, Correspondances, ColonnesPivot).
    """
    file_id = request.session.get("current_file_id")
    if not file_id:
        return JsonResponse({"success": False, "error": "Aucun fichier chargé"})
    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    if not filepath.exists():
        return JsonResponse({"success": False, "error": "Fichier introuvable"})
    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON invalide"})
    old_name = str(data.get("old_name", "")).strip()
    new_name = str(data.get("new_name", "")).strip()[:31]
    if not old_name or not new_name:
        return JsonResponse({"success": False, "error": "Ancien et nouveau noms requis"})
    if old_name == new_name:
        return JsonResponse({"success": True, "sheet_name": new_name})
    if old_name == CONFIG_SHEET or new_name == CONFIG_SHEET:
        return JsonResponse({"success": False, "error": "Impossible de renommer l'onglet _Config_"})
    try:
        wb = load_workbook(str(filepath))
        if old_name not in wb.sheetnames:
            return JsonResponse({"success": False, "error": f"Onglet « {old_name} » introuvable"})
        if new_name in wb.sheetnames:
            return JsonResponse({"success": False, "error": f"L'onglet « {new_name} » existe déjà"})
        ws = wb[old_name]
        ws.title = new_name
        config = read_config_from_excel(str(filepath))
        if config:
            for s in config.sheets:
                if s.name == old_name:
                    s.name = new_name
            for m in config.mappings:
                if m.source_sheet == old_name:
                    m.source_sheet = new_name
                if m.target_sheet == old_name:
                    m.target_sheet = new_name
            for p in config.pivot_columns:
                if p.sheet == old_name:
                    p.sheet = new_name
            wb.save(str(filepath))
            save_config_to_excel(config, str(filepath))
        else:
            wb.save(str(filepath))
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": True, "sheet_name": new_name})


@require_POST
def delete_sheet(request):
    """
    Supprime un onglet du fichier Excel.
    POST JSON: { sheet_name }
    Met à jour la config (supprime l'onglet, les correspondances et colonnes pivot associées).
    """
    file_id = request.session.get("current_file_id")
    if not file_id:
        return JsonResponse({"success": False, "error": "Aucun fichier chargé"})
    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    if not filepath.exists():
        return JsonResponse({"success": False, "error": "Fichier introuvable"})
    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON invalide"})
    sheet_name = str(data.get("sheet_name", "")).strip()
    if not sheet_name:
        return JsonResponse({"success": False, "error": "Nom d'onglet requis"})
    if sheet_name == CONFIG_SHEET:
        return JsonResponse({"success": False, "error": "Impossible de supprimer l'onglet _Config_"})
    try:
        config = read_config_from_excel(str(filepath))
        if config:
            config.sheets = [s for s in config.sheets if s.name != sheet_name]
            config.mappings = [m for m in config.mappings if m.source_sheet != sheet_name and m.target_sheet != sheet_name]
            config.pivot_columns = [p for p in config.pivot_columns if p.sheet != sheet_name]
        wb = load_workbook(str(filepath))
        if sheet_name not in wb.sheetnames:
            return JsonResponse({"success": False, "error": f"Onglet « {sheet_name} » introuvable"})
        wb.remove(wb[sheet_name])
        wb.save(str(filepath))
        if config:
            save_config_to_excel(config, str(filepath))
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": True, "message": f"Onglet « {sheet_name} » supprimé"})


@require_POST
def reset_config(request):
    """
    Vide l'arborescence existante et la configuration.
    Supprime tous les onglets sauf _Config_, puis réinitialise _Config_ avec des sections vides.
    """
    file_id = request.session.get("current_file_id")
    if not file_id:
        return JsonResponse({"success": False, "error": "Aucun fichier chargé"})
    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    if not filepath.exists():
        return JsonResponse({"success": False, "error": "Fichier introuvable"})
    try:
        wb = load_workbook(str(filepath))
        if CONFIG_SHEET not in wb.sheetnames:
            wb.create_sheet(CONFIG_SHEET, 0)
        to_remove = [sn for sn in wb.sheetnames if sn != CONFIG_SHEET]
        for sn in to_remove:
            wb.remove(wb[sn])
        config = ImportConfig()
        wb.save(str(filepath))
        save_config_to_excel(config, str(filepath))
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": True, "message": "Configuration réinitialisée"})


@require_GET
def export_config(request):
    """
    Exporte la configuration et les onglets sous Configuration vers un fichier Excel téléchargeable.
    """
    file_id = request.session.get("current_file_id")
    if not file_id:
        messages.error(request, "Aucun fichier chargé.")
        return redirect("explorer")
    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    if not filepath.exists():
        messages.error(request, "Fichier introuvable.")
        return redirect("explorer")
    try:
        with open(filepath, "rb") as f:
            content = f.read()
    except Exception as e:
        messages.error(request, str(e))
        return redirect("explorer")
    file_name = request.session.get("current_file_name", "import") or "import"
    if not file_name.lower().endswith(".xlsx"):
        file_name = file_name.rsplit(".", 1)[0] + ".xlsx" if "." in file_name else file_name + ".xlsx"
    base = file_name.rsplit(".", 1)[0]
    export_name = f"{base}_export.xlsx"
    response = HttpResponse(content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{export_name}"'
    return response


@require_POST
def run_import(request):
    """Lance l'import Salesforce."""
    file_id = request.session.get("current_file_id")
    if not file_id:
        return JsonResponse({"success": False, "error": "Aucun fichier chargé"})

    sf_client = _get_sf_client(request)
    if not sf_client:
        return JsonResponse({"success": False, "error": "Non connecté à Salesforce"})

    filepath = settings.MEDIA_ROOT / f"{file_id}.xlsx"
    if not filepath.exists():
        return JsonResponse({"success": False, "error": "Fichier introuvable"})

    try:
        data = read_excel_file(str(filepath))
        config = read_config_from_excel(str(filepath))
        if not config:
            config = create_default_config(list(data.keys()))
        data = add_import_columns(data, config)
        sheets_ordered = config.get_sheets_ordered()

        # Surcharger mode et champ upsert depuis le formulaire (interface)
        for s in sheets_ordered:
            post_mode = request.POST.get(f"mode_{s.name}", "").strip().lower()
            if post_mode in ("insert", "update", "upsert"):
                s.mode = post_mode
            post_upsert = request.POST.get(f"upsert_field_{s.name}", "")
            if post_upsert is not None:
                s.upsert_external_id_field = post_upsert.strip() or None

        for s in sheets_ordered:
            if not s.salesforce_object or not s.salesforce_object.strip():
                return JsonResponse({
                    "success": False,
                    "error": f"Objet Salesforce non défini pour l'onglet '{s.name}'",
                })
            s_mode = getattr(s, "mode", "insert") or "insert"
            if s_mode == "upsert" and not (getattr(s, "upsert_external_id_field", None) or "").strip():
                return JsonResponse({
                    "success": False,
                    "error": f"Mode Upsert pour l'onglet '{s.name}' : indiquez le champ externe ID (ex: RefExterne__c).",
                })

        logs = []
        failed_records = []
        results_by_sheet = {}
        current_data = {k: v.copy() for k, v in data.items()}

        for idx, sheet_config in enumerate(sheets_ordered):
            name = sheet_config.name
            logs.append(f"--- Onglet {idx + 1}/{len(sheets_ordered)}: {name} ---")
            df = current_data.get(name)
            if df is None or df.empty:
                logs.append("  Onglet vide, ignoré.")
                continue

            # Déterminer les lignes à ignorer (parent échoué) pour les onglets enfants
            skip_mask = pd.Series([False] * len(df), index=df.index)
            if idx > 0:
                for mapping in config.get_mappings_for_target(name):
                    src_results = results_by_sheet.get(mapping.source_sheet)
                    src_df = current_data.get(mapping.source_sheet)
                    if src_results is None or src_df is None or mapping.target_column not in df.columns:
                        continue
                    failed_keys = set()
                    for j, res in enumerate(src_results):
                        if not res.get("success") and j < len(src_df) and mapping.source_column in src_df.columns:
                            val = src_df.iloc[j].get(mapping.source_column)
                            if pd.notna(val) and str(val).strip():
                                failed_keys.add(str(val).strip())
                    for i in range(len(df)):
                        pivot_val = df.iloc[i].get(mapping.target_column)
                        if pd.notna(pivot_val) and str(pivot_val).strip() in failed_keys:
                            skip_mask.iloc[i] = True

            # Ajouter les lignes ignorées (parent échoué) aux failed_records avec message vert
            for i in range(len(df)):
                if skip_mask.iloc[i]:
                    for mapping in config.get_mappings_for_target(name):
                        src_results = results_by_sheet.get(mapping.source_sheet)
                        src_df = current_data.get(mapping.source_sheet)
                        if src_results is None or src_df is None:
                            continue
                        pivot_val = df.iloc[i].get(mapping.target_column)
                        if pd.notna(pivot_val) and str(pivot_val).strip():
                            failed_keys = set()
                            for j, res in enumerate(src_results):
                                if not res.get("success") and j < len(src_df) and mapping.source_column in src_df.columns:
                                    val = src_df.iloc[j].get(mapping.source_column)
                                    if pd.notna(val) and str(val).strip():
                                        failed_keys.add(str(val).strip())
                            if str(pivot_val).strip() in failed_keys:
                                row_data = {}
                                for col in df.columns:
                                    val = df.iloc[i].get(col)
                                    row_data[str(col)] = str(val)[:100] if pd.notna(val) else ""
                                failed_records.append({
                                    "sheet": name,
                                    "row_index": i + 1,
                                    "row_data": row_data,
                                    "error": f"Non importé : l'enregistrement parent ({mapping.source_sheet}.{mapping.source_column} = {pivot_val}) n'a pas pu être créé dans Salesforce.",
                                    "is_skipped_parent_failed": True,
                                })
                                break  # un seul message par ligne

            df_to_import = df[~skip_mask]
            if df_to_import.empty:
                logs.append("  Aucun enregistrement à importer (tous dépendent d'un parent en échec).")
                results_by_sheet[name] = [{"success": False, "id": None, "errors": []} for _ in range(len(df))]
                current_data = replicate_ids_to_next_sheets(current_data, config, results_by_sheet)
                continue

            # Colonnes pivot : utilisées pour la correspondance avec les onglets parents, non importées
            pivot_cols = config.get_pivot_columns_for_sheet(name)
            sheet_mode = getattr(sheet_config, "mode", "insert") or "insert"
            id_col = sheet_config.id_column or "Id"
            # Exclure Id en création, le garder pour update
            exclude_cols = list(pivot_cols)
            if sheet_mode == "insert" or sheet_mode == "upsert":
                exclude_cols.append(id_col)
            field_mapping = {col: col for col in df_to_import.columns}
            records = prepare_records_for_salesforce(df_to_import, field_mapping, exclude_columns=exclude_cols)
            if not records:
                logs.append("  Aucun enregistrement.")
                continue

            try:
                results = import_to_salesforce(
                    sf_client,
                    sheet_config.salesforce_object,
                    records,
                    mode=getattr(sheet_config, "mode", "insert") or "insert",
                    upsert_external_id_field=getattr(sheet_config, "upsert_external_id_field", None),
                )
            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)})

            # Aligner results avec df (insert None pour les lignes skippées)
            full_results = []
            result_idx = 0
            for i in range(len(df)):
                if skip_mask.iloc[i]:
                    full_results.append({"success": False, "id": None, "errors": []})
                else:
                    full_results.append(results[result_idx])
                    result_idx += 1
            results_by_sheet[name] = full_results

            success = sum(1 for r in full_results if r.get("success"))
            skipped = skip_mask.sum()
            if skipped:
                logs.append(f"  {success}/{len(results)} importés, {int(skipped)} ignoré(s) (parent en échec)")
            else:
                logs.append(f"  {success}/{len(full_results)} succès")

            # Collecter les lignes en échec (import Salesforce tenté mais a échoué) avec détails
            for j, res in enumerate(full_results):
                if not res.get("success") and not skip_mask.iloc[j]:
                    err_list = res.get("errors") or []
                    err_strs = []
                    for e in err_list:
                        if isinstance(e, dict):
                            err_strs.append(e.get("message", str(e)))
                        else:
                            err_strs.append(str(e))
                    err_msg = " ; ".join(err_strs) if err_strs else "Erreur inconnue"
                    row_data = {}
                    if j < len(df):
                        for col in df.columns:
                            val = df.iloc[j].get(col)
                            row_data[str(col)] = str(val)[:100] if pd.notna(val) else ""
                    failed_records.append({
                        "sheet": name,
                        "row_index": j + 1,
                        "row_data": row_data,
                        "error": err_msg,
                    })

            current_data = replicate_ids_to_next_sheets(current_data, config, results_by_sheet)

        logs.append("=== Import terminé ===")

        # Données mises à jour avec les IDs pour affichage
        sheets_data = []
        for sheet_config in sheets_ordered:
            name = sheet_config.name
            if name not in current_data:
                continue
            df = current_data[name]
            cols = list(df.columns)
            rows = []
            for _, row in df.iterrows():
                rows.append([str(v) if pd.notna(v) else "" for v in row])
            sheets_data.append({"name": name, "columns": cols, "rows": rows})

        return JsonResponse({
            "success": True,
            "logs": logs,
            "failed_records": failed_records,
            "sheets_data": sheets_data,
        })
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
