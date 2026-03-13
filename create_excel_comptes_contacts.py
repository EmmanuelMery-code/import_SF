"""
Script pour créer un fichier Excel d'exemple avec 3 onglets :
1. Comptes : 3 comptes parent
2. Contacts : 6 contacts (2 par compte), liés aux comptes via AccountId
3. SousComptes : 3 comptes enfants, liés aux comptes parent via ParentId
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_FILE = Path(__file__).parent / "exemple_comptes_contacts_souscomptes.xlsx"


def create_excel():
    wb = Workbook()
    ws_config = wb.active
    ws_config.title = "_Config_"

    # === SECTION 1 : Ordre des onglets ===
    ws_config.append(["OrdreOnglets"])
    ws_config.append(["", "NomOnglet", "Ordre", "ObjetSalesforce", "ColonneID"])
    ws_config.append(["", "Comptes", 1, "Account", "Id"])
    ws_config.append(["", "Contacts", 2, "Contact", "Id"])
    ws_config.append(["", "SousComptes", 3, "Account", "Id"])
    ws_config.append([])

    # === SECTION 2 : Correspondances ===
    ws_config.append(["Correspondances"])
    ws_config.append(["", "FeuilleSource", "ColonneSource", "FeuilleCible", "ColonneCible", "ChampSF"])
    # Comptes -> Contacts : RefCompte = RefExterne -> AccountId
    ws_config.append(["", "Comptes", "RefExterne", "Contacts", "RefCompte", "AccountId"])
    # Comptes -> SousComptes : RefParent = RefExterne -> ParentId
    ws_config.append(["", "Comptes", "RefExterne", "SousComptes", "RefParent", "ParentId"])
    ws_config.append([])

    # === SECTION 3 : Colonnes pivot (non importées) ===
    ws_config.append(["ColonnesPivot"])
    ws_config.append(["", "Feuille", "Colonne"])
    ws_config.append(["", "Contacts", "RefCompte"])
    ws_config.append(["", "SousComptes", "RefParent"])

    # === Feuille 1 : Comptes (3 comptes parent) ===
    ws_comptes = wb.create_sheet("Comptes")
    comptes = pd.DataFrame({
        "RefExterne": ["C001", "C002", "C003"],
        "Name": ["Entreprise Alpha", "Entreprise Beta", "Entreprise Gamma"],
        "BillingCity": ["Paris", "Lyon", "Marseille"],
    })
    for r in dataframe_to_rows(comptes, index=False, header=True):
        ws_comptes.append(r)

    # === Feuille 2 : Contacts (6 contacts, 2 par compte) ===
    ws_contacts = wb.create_sheet("Contacts")
    contacts = pd.DataFrame({
        "RefCompte": ["C001", "C001", "C002", "C002", "C003", "C003"],
        "LastName": ["Dupont", "Martin", "Bernard", "Petit", "Leroy", "Moreau"],
        "FirstName": ["Jean", "Marie", "Pierre", "Sophie", "Luc", "Claire"],
        "Email": [
            "jean.dupont@alpha.com",
            "marie.martin@alpha.com",
            "pierre.bernard@beta.com",
            "sophie.petit@beta.com",
            "luc.leroy@gamma.com",
            "claire.moreau@gamma.com",
        ],
    })
    for r in dataframe_to_rows(contacts, index=False, header=True):
        ws_contacts.append(r)

    # === Feuille 3 : SousComptes (3 comptes enfants, ParentId = Id d'un compte du 1er onglet) ===
    ws_sous = wb.create_sheet("SousComptes")
    sous_comptes = pd.DataFrame({
        "RefParent": ["C001", "C002", "C003"],
        "Name": ["Filiale Alpha Nord", "Filiale Beta Sud", "Filiale Gamma Ouest"],
        "BillingCity": ["Lille", "Nice", "Nantes"],
    })
    for r in dataframe_to_rows(sous_comptes, index=False, header=True):
        ws_sous.append(r)

    wb.save(OUTPUT_FILE)
    print(f"Fichier créé : {OUTPUT_FILE}")
    print("\nStructure : 3 onglets")
    print("  1. Comptes : 3 comptes parent (RefExterne: C001, C002, C003)")
    print("  2. Contacts : 6 contacts (2 par compte), liés via RefCompte -> AccountId")
    print("  3. SousComptes : 3 comptes enfants, ParentId = Id d'un compte du 1er onglet (RefParent -> ParentId)")


if __name__ == "__main__":
    create_excel()
